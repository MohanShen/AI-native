"""
Excelæ–‡ä»¶é¢„å¤„ç†æ¨¡å—
å°†å¤æ‚Excelè¡¨æ ¼é‡å¡‘ä¸ºäºŒç»´è¡¨ï¼Œæ”¯æŒå¤æ‚è¡¨å¤´ç»“æ„
"""
import hashlib
import json
import logging
import os
import uuid
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter

# é…ç½®æ—¥å¿—
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelPreprocessor:
    """Excelæ–‡ä»¶é¢„å¤„ç†å™¨ï¼Œæ”¯æŒå¤æ‚è¡¨å¤´ç»“æ„"""
    
    def __init__(self, knowledge_base_path: str = "knowledge_base", openai_client=None, use_llm_analysis: bool = True):
        """
        åˆå§‹åŒ–é¢„å¤„ç†å™¨
        
        Args:
            knowledge_base_path: çŸ¥è¯†åº“è·¯å¾„ï¼ŒåŒ…å«Excelæ–‡ä»¶
            openai_client: OpenAIå®¢æˆ·ç«¯ï¼ˆç”¨äºLLMåˆ†æè¡¨å¤´ç»“æ„ï¼‰
            use_llm_analysis: æ˜¯å¦ä½¿ç”¨LLMåˆ†æå¤æ‚è¡¨å¤´ï¼ˆé»˜è®¤Trueï¼‰
        """
        self.knowledge_base_path = knowledge_base_path
        self.openai_client = openai_client
        self.use_llm_analysis = use_llm_analysis
        self.processed_files: Dict[str, pd.DataFrame] = {}
        self.file_metadata: Dict[str, Dict] = {}
        
        # ä¸´æ—¶ç›®å½•ç”¨äºå­˜å‚¨é‡å»ºçš„æ–‡ä»¶
        self.temp_dir = Path(knowledge_base_path) / ".reconstructed"
        self.temp_dir.mkdir(exist_ok=True)
        
        # ç¡®ä¿çŸ¥è¯†åº“ç›®å½•å­˜åœ¨
        os.makedirs(knowledge_base_path, exist_ok=True)
    
    def load_excel_file(self, file_path: str) -> pd.DataFrame:
        """
        åŠ è½½Excelæ–‡ä»¶å¹¶é‡å¡‘ä¸ºäºŒç»´è¡¨
        æ”¯æŒå¤æ‚è¡¨å¤´ç»“æ„çš„å¤„ç†
        
        Args:
            file_path: Excelæ–‡ä»¶è·¯å¾„
            
        Returns:
            å¤„ç†åçš„DataFrame
        """
        try:
            file_name = os.path.basename(file_path)
            
            # å®é™…ç”¨äºåˆ†æçš„æ•°æ®æ–‡ä»¶è·¯å¾„ï¼ˆé»˜è®¤æ˜¯åŸå§‹æ–‡ä»¶ï¼‰
            actual_data_path = file_path
            
            # å¦‚æœä½¿ç”¨LLMåˆ†æä¸”OpenAIå®¢æˆ·ç«¯å¯ç”¨ï¼Œä½¿ç”¨å¤æ‚å¤„ç†æµç¨‹
            if self.use_llm_analysis and self.openai_client:
                try:
                    logger.info(f"ä½¿ç”¨LLMåˆ†æå¤„ç†æ–‡ä»¶: {file_name}")
                    processed_file = self._process_with_llm(file_path)
                    if processed_file and os.path.exists(processed_file):
                        # ä»å¤„ç†åçš„æ–‡ä»¶è¯»å–
                        df = self._read_processed_file(processed_file)
                        # å¯¹äºåç»­åˆ†æå’Œä»£ç æ‰§è¡Œï¼Œåº”ä½¿ç”¨é‡å»ºåçš„æ–‡ä»¶è·¯å¾„
                        actual_data_path = processed_file
                    else:
                        # å›é€€åˆ°ç®€å•å¤„ç†
                        logger.warning(f"LLMå¤„ç†å¤±è´¥ï¼Œä½¿ç”¨ç®€å•å¤„ç†: {file_name}")
                        df = self._simple_load(file_path)
                except Exception as e:
                    logger.warning(f"LLMå¤„ç†å‡ºé”™ï¼Œä½¿ç”¨ç®€å•å¤„ç†: {file_name}, é”™è¯¯: {str(e)}")
                    df = self._simple_load(file_path)
            else:
                # ä½¿ç”¨ç®€å•å¤„ç†
                df = self._simple_load(file_path)
            
            # é‡å¡‘ä¸ºäºŒç»´è¡¨ï¼šæ¸…ç†ç©ºè¡Œç©ºåˆ—ï¼Œé‡ç½®ç´¢å¼•
            df = self._reshape_to_2d(df)
            
            # æ‰“å°é¢„å¤„ç†åçš„è¡¨å¤´ä¿¡æ¯
            self._print_preprocessed_headers(file_name, df)
            
            # å­˜å‚¨å¤„ç†åçš„æ•°æ®
            self.processed_files[file_name] = df
            # å°†å®é™…ç”¨äºåˆ†æçš„æ•°æ®æ–‡ä»¶è·¯å¾„æ”¾å…¥DataFrameå±æ€§ï¼Œä½œä¸ºé¢å¤–ä¿é™©
            df.attrs['file_path'] = actual_data_path
            
            # å­˜å‚¨å…ƒæ•°æ®
            self.file_metadata[file_name] = {
                # è¿™é‡Œçš„pathç”¨äºåç»­ä»£ç æ‰§è¡Œï¼Œåº”æŒ‡å‘å®é™…çš„æ•°æ®æ–‡ä»¶ï¼ˆé‡å»ºåæ–‡ä»¶ï¼‰
                'path': actual_data_path,
                'columns': list(df.columns),
                'shape': df.shape,
                'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()}
            }
            
            return df
            
        except Exception as e:
            raise Exception(f"åŠ è½½Excelæ–‡ä»¶å¤±è´¥: {str(e)}")
    
    def _simple_load(self, file_path: str) -> pd.DataFrame:
        """
        ç®€å•åŠ è½½æ–¹æ³•ï¼ˆåŸæœ‰é€»è¾‘ï¼‰
        
        Args:
            file_path: Excelæ–‡ä»¶è·¯å¾„
            
        Returns:
            DataFrame
        """
        excel_file = pd.ExcelFile(file_path)
        sheets = excel_file.sheet_names
        
        # å¦‚æœåªæœ‰ä¸€ä¸ªsheetï¼Œç›´æ¥è¯»å–ï¼ˆheader=0è¡¨ç¤ºç¬¬ä¸€è¡Œä½œä¸ºåˆ—åï¼‰
        if len(sheets) == 1:
            df = pd.read_excel(file_path, sheet_name=sheets[0], header=0)
        else:
            # å¤šä¸ªsheetæ—¶ï¼Œå°è¯•æ‰¾åˆ°æ•°æ®æœ€å¤šçš„sheet
            max_rows = 0
            best_sheet = sheets[0]
            for sheet in sheets:
                temp_df = pd.read_excel(file_path, sheet_name=sheet, header=0)
                if len(temp_df) > max_rows:
                    max_rows = len(temp_df)
                    best_sheet = sheet
            df = pd.read_excel(file_path, sheet_name=best_sheet, header=0)
        
        return df
    
    def _process_with_llm(self, file_path: str) -> Optional[str]:
        """
        ä½¿ç”¨LLMå¤„ç†å¤æ‚è¡¨å¤´ç»“æ„
        
        Args:
            file_path: åŸå§‹Excelæ–‡ä»¶è·¯å¾„
            
        Returns:
            å¤„ç†åçš„æ–‡ä»¶è·¯å¾„ï¼Œå¦‚æœå¤±è´¥è¿”å›None
        """
        try:
            # æ£€æŸ¥æ˜¯å¦å·²æœ‰å¤„ç†åçš„æ–‡ä»¶
            existing_recon = self._get_reconstructed_path(file_path)
            if existing_recon and os.path.exists(existing_recon):
                # æ£€æŸ¥åŸå§‹æ–‡ä»¶æ˜¯å¦å·²ä¿®æ”¹
                original_stat = Path(file_path).stat()
                recon_stat = Path(existing_recon).stat()
                
                if original_stat.st_mtime <= recon_stat.st_mtime:
                    logger.info(f"ä½¿ç”¨å·²å­˜åœ¨çš„é‡å»ºæ–‡ä»¶: {existing_recon}")
                    return existing_recon
                else:
                    logger.info(f"åŸå§‹æ–‡ä»¶å·²ä¿®æ”¹ï¼Œé‡æ–°ç”Ÿæˆ...")
                    os.remove(existing_recon)
            
            # ç”Ÿæˆè¾“å‡ºè·¯å¾„
            original_name = Path(file_path).stem
            file_hash = hashlib.md5(str(Path(file_path).absolute()).encode()).hexdigest()[:8]
            output_path = self.temp_dir / f"{original_name}_reconstructed_{file_hash}.xlsx"
            
            logger.info(f"å¤„ç†Excelæ–‡ä»¶: {file_path}")
            logger.info(f"é‡å»ºæ–‡ä»¶å°†ä¿å­˜åˆ°: {output_path}")
            
            # Step 1: å–æ¶ˆåˆå¹¶å•å…ƒæ ¼å¹¶å¡«å……
            unmerged_file, merged_info = self._step1_unmerge_and_fill(file_path)
            
            try:
                # Step 2: LLMåˆ†æè¡¨å¤´ç»“æ„
                analysis_result = self._step2_model_analysis(unmerged_file, merged_info)
                
                # Step 3: è‡ªåŠ¨å¤„ç†ï¼ˆåˆ é™¤æ ‡ç­¾è¡Œï¼Œåˆå¹¶å¤šçº§è¡¨å¤´ï¼‰
                reconstructed_data = self._step3_automated_processing(unmerged_file, analysis_result)
                
                # å†™å…¥é‡å»ºæ–‡ä»¶
                self._write_reconstructed_file(reconstructed_data, output_path)
                
                logger.info(f"Excelå¤„ç†å®Œæˆã€‚é‡å»ºæ–‡ä»¶: {output_path}")
                return str(output_path)
                
            finally:
                # æ¸…ç†ä¸´æ—¶æ–‡ä»¶
                if os.path.exists(unmerged_file):
                    os.remove(unmerged_file)
                    
        except Exception as e:
            logger.error(f"LLMå¤„ç†Excelæ–‡ä»¶æ—¶å‡ºé”™: {e}", exc_info=True)
            return None
    
    def _step1_unmerge_and_fill(self, file_path: str) -> Tuple[str, Dict]:
        """
        Step 1: é¢„å¤„ç† - å–æ¶ˆåˆå¹¶å•å…ƒæ ¼å¹¶å¡«å……ç©ºç™½
        
        Args:
            file_path: åŸå§‹Excelæ–‡ä»¶è·¯å¾„
            
        Returns:
            (å–æ¶ˆåˆå¹¶åçš„æ–‡ä»¶è·¯å¾„, åˆå¹¶ä¿¡æ¯å­—å…¸)
        """
        try:
            logger.info("Step 1: å–æ¶ˆåˆå¹¶å•å…ƒæ ¼å¹¶å¡«å……ç©ºç™½...")
            
            # åˆ›å»ºä¸´æ—¶æ–‡ä»¶
            unmerged_file = str(self.temp_dir / f"unmerged_{uuid.uuid4().hex[:8]}.xlsx")
            
            # åŠ è½½å·¥ä½œç°¿
            wb = openpyxl.load_workbook(file_path, data_only=True)
            merged_info = {}
            
            for ws in wb.worksheets:
                logger.info(f"  å¤„ç†å·¥ä½œè¡¨: {ws.title}")
                sheet_merged_info = []
                
                # æ”¶é›†åˆå¹¶ä¿¡æ¯å¹¶å–æ¶ˆåˆå¹¶
                for merged_range in list(ws.merged_cells.ranges):
                    min_row, min_col, max_row, max_col = (
                        merged_range.min_row, merged_range.min_col,
                        merged_range.max_row, merged_range.max_col
                    )
                    value = ws.cell(row=min_row, column=min_col).value
                    
                    # å­˜å‚¨åˆå¹¶ä¿¡æ¯
                    sheet_merged_info.append({
                        "range": str(merged_range),
                        "start": (min_row, min_col),
                        "end": (max_row, max_col),
                        "value": value
                    })
                    
                    # å–æ¶ˆåˆå¹¶
                    ws.unmerge_cells(start_row=min_row, start_column=min_col,
                                   end_row=max_row, end_column=max_col)
                    
                    # å¡«å……æ‰€æœ‰å•å…ƒæ ¼
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            ws.cell(row=row, column=col, value=value)
                
                merged_info[ws.title] = sheet_merged_info
                logger.info(f"    å–æ¶ˆåˆå¹¶äº† {len(sheet_merged_info)} ä¸ªåˆå¹¶å•å…ƒæ ¼èŒƒå›´")
            
            # ä¿å­˜å–æ¶ˆåˆå¹¶åçš„æ–‡ä»¶
            wb.save(unmerged_file)
            wb.close()
            logger.info(f"Step 1 å®Œæˆã€‚å–æ¶ˆåˆå¹¶æ–‡ä»¶: {os.path.basename(unmerged_file)}")
            
            return unmerged_file, merged_info
            
        except Exception as e:
            logger.error(f"Step 1 (å–æ¶ˆåˆå¹¶) å‡ºé”™: {e}", exc_info=True)
            raise
    
    def _step2_model_analysis(self, unmerged_file: str, merged_info: Dict) -> List[Dict]:
        """
        Step 2: æ¨¡å‹åˆ†æ - è¯†åˆ«è¡¨å¤´å’Œæ ‡ç­¾è¡Œ
        
        Args:
            unmerged_file: å–æ¶ˆåˆå¹¶åçš„æ–‡ä»¶è·¯å¾„
            merged_info: åˆå¹¶ä¿¡æ¯å­—å…¸
            
        Returns:
            åˆ†æç»“æœåˆ—è¡¨
        """
        if self.openai_client is None:
            logger.warning("æœªæä¾›OpenAIå®¢æˆ·ç«¯ï¼Œä½¿ç”¨é»˜è®¤åˆ†æ")
            all_sheets = pd.read_excel(unmerged_file, sheet_name=None, header=None)
            return [
                {sheet_name: {"labels": [], "header": [1]}}
                for sheet_name in all_sheets.keys()
            ]
        
        try:
            logger.info("Step 2: æ¨¡å‹åˆ†æ - è¯†åˆ«æ ‡ç­¾è¡Œå’Œè¡¨å¤´...")
            
            # æå–å‰10è¡Œä½œä¸ºæ ·æœ¬
            excel_info = self._get_excel_data(unmerged_file, head=10)
            
            # å‡†å¤‡åˆå¹¶ä¿¡æ¯
            merged_info_json = json.dumps(merged_info, ensure_ascii=False, indent=2)
            
            system_prompt = '''ä½ æ˜¯ä¸€ä¸ªä¸“ä¸šçš„ç»“æ„åŒ–æ•°æ®å¤„ç†AIï¼Œä¸“é—¨åˆ†æExcelè¡¨æ ¼ç»“æ„ã€‚

æ ¸å¿ƒèƒ½åŠ›ï¼š
1. å‡†ç¡®è¯†åˆ«è¡¨å¤´è¡Œï¼šè¡¨å¤´é€šå¸¸æ˜¯ç®€çŸ­çš„æ ‡ç­¾ï¼ˆ1-5ä¸ªè¯ï¼‰ï¼Œæè¿°æ•°æ®åˆ—ã€‚å®ƒä»¬å‡ºç°åœ¨è¡¨æ ¼é¡¶éƒ¨ï¼Œç»“æ„ä¸€è‡´ã€‚
2. åŒºåˆ†è¡¨å¤´å’Œæ•°æ®ï¼šæ•°æ®è¡ŒåŒ…å«å®é™…å€¼ï¼ˆæ•°å­—ã€æ—¥æœŸã€é•¿æè¿°ã€å¸¦è¯¦ç»†ä¿¡æ¯çš„äº§å“åç§°ï¼‰ã€‚è¡¨å¤´æ˜¯ç®€æ´çš„åˆ—æ ‡ç­¾ã€‚
3. è¯†åˆ«å¤šçº§è¡¨å¤´ï¼šåªæœ‰é¡¶éƒ¨è¿ç»­çš„è¡Œï¼Œæ˜æ˜¾æ˜¯è¡¨å¤´æ ‡ç­¾ï¼ˆä¸æ˜¯æ•°æ®ï¼‰ï¼Œæ‰åº”è¢«è§†ä¸ºè¡¨å¤´ã€‚
4. è¯†åˆ«æ ‡ç­¾è¡Œï¼šå·¥ä½œè¡¨çº§åˆ«çš„æè¿°ã€æ ‡é¢˜æˆ–æ³¨é‡Šï¼Œå‡ºç°åœ¨å®é™…æ•°æ®è¡¨ä¹‹å‰ï¼ˆä¸æ˜¯æ•°æ®è¡Œå†…å®¹ï¼‰ã€‚

å…³é”®è§„åˆ™ï¼š
- è¡¨å¤´æ˜¯ç®€çŸ­çš„æ ‡ç­¾ï¼ˆé€šå¸¸æ¯ä¸ªå•å…ƒæ ¼1-5ä¸ªè¯ï¼‰ï¼Œä¸æ˜¯é•¿æè¿°æˆ–æ•°æ®å€¼
- å¦‚æœä¸€è¡ŒåŒ…å«é•¿æ–‡æœ¬ã€æ•°å­—æˆ–è¯¦ç»†çš„äº§å“ä¿¡æ¯ï¼Œå®ƒæ˜¯æ•°æ®ï¼Œä¸æ˜¯è¡¨å¤´
- å¤šçº§è¡¨å¤´é€šå¸¸æœ€å¤š1-3è¡Œï¼Œéƒ½åœ¨æœ€é¡¶éƒ¨
- æ•°æ®è¡Œç»ä¸èƒ½è¢«è§†ä¸ºè¡¨å¤´'''
            
            user_prompt = f'''è¯·åˆ†ææ¯ä¸ªå·¥ä½œè¡¨çš„ç»“æ„å¹¶è¯†åˆ«ï¼š
1. æ ‡ç­¾è¡Œï¼šè¦åˆ é™¤çš„å·¥ä½œè¡¨çº§åˆ«æ ‡é¢˜/æ³¨é‡Š/æè¿°ï¼ˆåœ¨å®é™…è¡¨æ ¼ä¹‹å‰ï¼‰
2. è¡¨å¤´è¡Œï¼šå®é™…çš„åˆ—è¡¨å¤´è¡Œ - è¿™äº›åº”è¯¥æ˜¯ç®€çŸ­çš„æ ‡ç­¾ï¼Œä¸æ˜¯æ•°æ®

é‡è¦ï¼šè¡¨å¤´æ˜¯ç®€æ´çš„åˆ—æ ‡ç­¾ã€‚å¦‚æœä¸€è¡ŒåŒ…å«é•¿æè¿°ã€äº§å“è¯¦ç»†ä¿¡æ¯æˆ–å®é™…æ•°æ®å€¼ï¼Œå®ƒæ˜¯æ•°æ®è¡Œï¼Œä¸æ˜¯è¡¨å¤´ã€‚

è¦åˆ†æçš„æ•°æ®ï¼š

1. å–æ¶ˆåˆå¹¶åçš„Excelæ–‡ä»¶æ•°æ®ï¼ˆå‰10è¡Œï¼‰ï¼š

```
{excel_info}
```

2. åŸå§‹åˆå¹¶å•å…ƒæ ¼ä¿¡æ¯ï¼ˆç”¨äºç¡®å®šè¡¨å¤´çº§åˆ«ï¼‰ï¼š

```
{merged_info_json}
```

è¾“å‡ºæ ¼å¼ï¼š

[
    {{
        "sheet_name1": {{
            "labels": [è¡Œå·],    # æ•´ä¸ªå·¥ä½œè¡¨çš„æ ‡ç­¾æ–‡æœ¬è¡Œï¼ˆå¦‚æœæ²¡æœ‰åˆ™ä¸ºç©ºåˆ—è¡¨ï¼‰
            "header": [è¡Œå·]      # å¤šçº§è¡¨å¤´è¡Œï¼ˆå¿…é¡»åŒ…å«è‡³å°‘1è¡Œï¼‰
        }},
        "sheet_name2": {{
            "labels": [è¡Œå·],
            "header": [è¡Œå·]
        }}
    }}
]

å…³é”®æŒ‡å¯¼åŸåˆ™ï¼š
1. è¡¨å¤´æ˜¯ç®€çŸ­çš„æ ‡ç­¾ï¼ˆæ¯ä¸ªå•å…ƒæ ¼1-5ä¸ªè¯ï¼‰ã€‚é•¿æ–‡æœ¬ = æ•°æ®è¡Œï¼Œä¸æ˜¯è¡¨å¤´ã€‚
2. è¡¨å¤´é€šå¸¸å‡ºç°åœ¨å‰1-3è¡Œã€‚å¦‚æœçœ‹åˆ°å®é™…æ•°æ®å€¼ï¼ˆæ•°å­—ã€å¸¦è¯¦ç»†ä¿¡æ¯çš„äº§å“åç§°ï¼‰ï¼Œé‚£æ˜¯æ•°æ®è¡Œã€‚
3. å¤šçº§è¡¨å¤´å¾ˆå°‘è§ - é€šå¸¸åªæœ‰1-2è¡Œã€‚åªæœ‰å½“å¤šè¡Œæ˜æ˜¾éƒ½æ˜¯è¡¨å¤´æ ‡ç­¾ï¼ˆç®€çŸ­ã€æè¿°æ€§ï¼‰æ—¶ï¼Œæ‰æ ‡è®°å¤šè¡Œä¸ºè¡¨å¤´ã€‚
4. å¦‚æœ‰ç–‘é—®ï¼Œä½¿ç”¨æ›´å°‘çš„è¡¨å¤´è¡Œã€‚æœ‰1ä¸ªè¡¨å¤´è¡Œæ¯”å°†æ•°æ®åˆå¹¶åˆ°è¡¨å¤´ä¸­æ›´å¥½ã€‚
5. æ ‡ç­¾è¡Œæ˜¯è¡¨æ ¼ä¹‹å‰çš„å·¥ä½œè¡¨çº§åˆ«æè¿°/æ ‡é¢˜ï¼Œä¸æ˜¯æ•°æ®å†…å®¹ã€‚
6. æ¯ä¸ªå·¥ä½œè¡¨ç‹¬ç«‹åˆ†æã€‚
7. å·¥ä½œè¡¨åç§°å¿…é¡»å®Œå…¨åŒ¹é…ã€‚
8. åªè¾“å‡ºJSONç»“æœï¼Œä¸è¦è§£é‡Šã€‚

ç¤ºä¾‹æ­£ç¡®è¾“å‡ºï¼š

[
    {{"sheet_name1": {{
        "labels": [1, 2],
        "header": [3, 4, 5]
    }}}},
    {{"sheet_name2": {{
        "labels": [],
        "header": [1, 2]
    }}}},
    {{"sheet_name3": {{
        "labels": [],
        "header": [1]
    }}}}
]'''
            
            response = self.openai_client.chat.completions.create(
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.3
            )
            
            result_text = response.choices[0].message.content.strip()
            # æ¸…ç†JSONï¼ˆå¯èƒ½åŒ…å«markdownä»£ç å—ï¼‰
            result_text = result_text.replace('```json', '').replace('```', '').strip()
            
            analysis_result = json.loads(result_text)
            logger.info(f"Step 2 å®Œæˆã€‚åˆ†æäº† {len(analysis_result)} ä¸ªå·¥ä½œè¡¨")
            
            return analysis_result
            
        except Exception as e:
            logger.error(f"Step 2 (æ¨¡å‹åˆ†æ) å‡ºé”™: {e}", exc_info=True)
            logger.info("å›é€€åˆ°é»˜è®¤åˆ†æ")
            all_sheets = pd.read_excel(unmerged_file, sheet_name=None, header=None)
            return [
                {sheet_name: {"labels": [], "header": [1]}}
                for sheet_name in all_sheets.keys()
            ]
    
    def _step3_automated_processing(self, unmerged_file: str, analysis_result: List[Dict]) -> Dict:
        """
        Step 3: è‡ªåŠ¨å¤„ç† - æ¸…ç†å’Œåˆå¹¶
        
        Args:
            unmerged_file: å–æ¶ˆåˆå¹¶åçš„æ–‡ä»¶è·¯å¾„
            analysis_result: Step 2çš„åˆ†æç»“æœ
            
        Returns:
            é‡å»ºæ•°æ®çš„å­—å…¸ï¼ˆå·¥ä½œè¡¨å -> DataFrameï¼‰
        """
        try:
            logger.info("Step 3: è‡ªåŠ¨å¤„ç† - æ¸…ç†å’Œåˆå¹¶è¡¨å¤´...")
            
            reconstructed_data = {}
            
            for sheet_config in analysis_result:
                for sheet_name, config in sheet_config.items():
                    labels = config.get('labels', [])
                    header = config.get('header', [1])
                    
                    logger.info(f"  å¤„ç†å·¥ä½œè¡¨: {sheet_name}")
                    logger.info(f"    è¦åˆ é™¤çš„æ ‡ç­¾è¡Œ: {labels}")
                    logger.info(f"    è¡¨å¤´è¡Œ: {header}")
                    
                    # Step 3.1: å…ˆåˆ é™¤æ ‡ç­¾è¡Œï¼ˆåœ¨è¯»å–è¡¨å¤´ä¹‹å‰ï¼‰
                    df_raw = pd.read_excel(unmerged_file, sheet_name=sheet_name, header=None, dtype=object)
                    
                    if labels:
                        # è½¬æ¢ä¸º0åŸºç´¢å¼•å¹¶åˆ é™¤æ ‡ç­¾è¡Œ
                        labels_0_based = [x - 1 for x in labels]
                        df_raw = df_raw.drop(labels_0_based, axis=0, errors='ignore')
                        df_raw = df_raw.reset_index(drop=True)
                        logger.info(f"    åˆ é™¤äº† {len(labels)} ä¸ªæ ‡ç­¾è¡Œ")
                    
                    # Step 3.2: è°ƒæ•´è¡¨å¤´è¡Œå·ï¼ˆåˆ é™¤æ ‡ç­¾è¡Œåï¼‰
                    header_0_based = self._adjust_header_indices(header, labels, len(df_raw))
                    
                    if header_0_based is None:
                        # ç©ºæ•°æ®æ¡†æˆ–æ— æ•ˆè¡¨å¤´
                        reconstructed_data[sheet_name] = pd.DataFrame()
                        continue
                    
                    # Step 3.3: è®¾ç½®è¡¨å¤´è¡Œå¹¶æå–æ•°æ®
                    df = self._extract_data_with_headers(df_raw, header_0_based)
                    
                    # æ¸…ç†åˆ—åå’Œæ•°æ®
                    df.columns = self._clean_column_names(df.columns)
                    
                    if len(header_0_based) > 1:
                        logger.info(f"    å°† {len(header_0_based)} ä¸ªè¡¨å¤´è¡Œåˆå¹¶ä¸ºå•è¡Œ")
                    
                    # åˆ é™¤å®Œå…¨ç©ºçš„è¡Œ
                    df = df.dropna(how='all')
                    
                    reconstructed_data[sheet_name] = df
                    logger.info(f"    æœ€ç»ˆ: {len(df)} è¡Œ Ã— {len(df.columns)} åˆ—")
                    
                    # æ‰“å°é‡å»ºåçš„è¡¨å¤´ä¿¡æ¯
                    print(f"\nğŸ“‹ å·¥ä½œè¡¨ '{sheet_name}' é‡å»ºåçš„è¡¨å¤´:")
                    print(f"  åˆ—å: {list(df.columns)}")
                    if len(df) > 0:
                        print(f"  å‰3è¡Œæ•°æ®:")
                        print(df.head(3).to_string())
                        print()
            
            logger.info(f"Step 3 å®Œæˆã€‚å¤„ç†äº† {len(reconstructed_data)} ä¸ªå·¥ä½œè¡¨")
            return reconstructed_data
            
        except Exception as e:
            logger.error(f"Step 3 (è‡ªåŠ¨å¤„ç†) å‡ºé”™: {e}", exc_info=True)
            raise
    
    def _get_excel_data(self, file_path: str, head: int = 10) -> str:
        """
        è·å–Excelæ–‡ä»¶çš„æ ·æœ¬æ•°æ®ç”¨äºLLMåˆ†æ
        
        Args:
            file_path: Excelæ–‡ä»¶è·¯å¾„
            head: æ¯ä¸ªå·¥ä½œè¡¨æå–çš„è¡Œæ•°
            
        Returns:
            æ ¼å¼åŒ–çš„å­—ç¬¦ä¸²ï¼ŒåŒ…å«å·¥ä½œè¡¨ä¿¡æ¯
        """
        try:
            all_sheets_data = pd.read_excel(file_path, sheet_name=None, header=None)
            prompt_parts = []
            
            for sheet_name, data in all_sheets_data.items():
                data.index = data.index + 1  # 1åŸºç´¢å¼•
                excel_col_names = [get_column_letter(i + 1) for i in range(len(data.columns))]
                data.columns = excel_col_names
                
                # æ›¿æ¢å­—ç¬¦ä¸²ä¸­çš„æ¢è¡Œç¬¦
                data = data.map(lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x)
                
                # ä½¿ç”¨to_stringä»£æ›¿to_markdownï¼Œé¿å…ä¾èµ–tabulate
                try:
                    sheet_sample = data.head(head).to_markdown(index=True)
                except ImportError:
                    # å¦‚æœæ²¡æœ‰tabulateï¼Œä½¿ç”¨to_string
                    sheet_sample = data.head(head).to_string(index=True)
                
                sheet_info = f"å·¥ä½œè¡¨: {sheet_name}\nå‰ {head} è¡Œ:\n\n{sheet_sample}\n\n---"
                prompt_parts.append(sheet_info)
            
            return '\n'.join(prompt_parts)
            
        except Exception as e:
            logger.error(f"æå–Excelæ•°æ®æ—¶å‡ºé”™: {e}", exc_info=True)
            raise
    
    def _write_reconstructed_file(self, reconstructed_data: Dict, output_path: Path) -> None:
        """
        å°†é‡å»ºæ•°æ®å†™å…¥Excelæ–‡ä»¶
        
        Args:
            reconstructed_data: å·¥ä½œè¡¨å -> DataFrame çš„å­—å…¸
            output_path: è¾“å‡ºæ–‡ä»¶è·¯å¾„
        """
        try:
            logger.info(f"å†™å…¥é‡å»ºæ–‡ä»¶: {output_path}")
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in reconstructed_data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    logger.info(f"  - å·¥ä½œè¡¨ '{sheet_name}': {len(df)} è¡Œ Ã— {len(df.columns)} åˆ—")
            
            logger.info(f"é‡å»ºæ–‡ä»¶ä¿å­˜æˆåŠŸ")
            
        except Exception as e:
            logger.error(f"å†™å…¥é‡å»ºæ–‡ä»¶æ—¶å‡ºé”™: {e}", exc_info=True)
            raise
    
    def _get_reconstructed_path(self, original_path: str) -> Optional[str]:
        """
        è·å–é‡å»ºæ–‡ä»¶è·¯å¾„ï¼ˆå¦‚æœå­˜åœ¨ï¼‰
        
        Args:
            original_path: åŸå§‹æ–‡ä»¶è·¯å¾„
            
        Returns:
            é‡å»ºæ–‡ä»¶è·¯å¾„ï¼Œå¦‚æœä¸å­˜åœ¨åˆ™è¿”å›None
        """
        original_name = Path(original_path).stem
        
        # æŸ¥æ‰¾ç°æœ‰çš„é‡å»ºæ–‡ä»¶
        for file in self.temp_dir.glob(f"{original_name}_reconstructed_*.xlsx"):
            return str(file)
        
        return None
    
    def _read_processed_file(self, processed_file: str) -> pd.DataFrame:
        """
        ä»å¤„ç†åçš„æ–‡ä»¶è¯»å–æ•°æ®
        
        Args:
            processed_file: å¤„ç†åçš„æ–‡ä»¶è·¯å¾„
            
        Returns:
            DataFrame
        """
        excel_file = pd.ExcelFile(processed_file)
        sheets = excel_file.sheet_names
        
        # é€‰æ‹©æ•°æ®æœ€å¤šçš„sheet
        if len(sheets) == 1:
            return pd.read_excel(processed_file, sheet_name=sheets[0], header=0)
        else:
            max_rows = 0
            best_sheet = sheets[0]
            for sheet in sheets:
                temp_df = pd.read_excel(processed_file, sheet_name=sheet, header=0)
                if len(temp_df) > max_rows:
                    max_rows = len(temp_df)
                    best_sheet = sheet
            return pd.read_excel(processed_file, sheet_name=best_sheet, header=0)
    
    def _adjust_header_indices(self, header: List[int], labels: List[int], df_length: int) -> Optional[List[int]]:
        """
        è°ƒæ•´è¡¨å¤´è¡Œç´¢å¼•ï¼ˆåˆ é™¤æ ‡ç­¾è¡Œåï¼‰
        
        Args:
            header: åŸå§‹è¡¨å¤´è¡Œå·ï¼ˆ1åŸºï¼‰
            labels: å·²åˆ é™¤çš„æ ‡ç­¾è¡Œå·ï¼ˆ1åŸºï¼‰
            df_length: åˆ é™¤æ ‡ç­¾è¡Œåçš„æ•°æ®æ¡†é•¿åº¦
            
        Returns:
            è°ƒæ•´åçš„è¡¨å¤´ç´¢å¼•ï¼ˆ0åŸºï¼‰ï¼Œå¦‚æœæ— æ•ˆåˆ™è¿”å›None
        """
        if df_length == 0:
            return None
        
        # è°ƒæ•´è¡¨å¤´ç´¢å¼•ï¼ˆè€ƒè™‘å·²åˆ é™¤çš„æ ‡ç­¾è¡Œï¼‰
        if labels:
            header_adjusted = [h - sum(1 for l in labels if l < h) for h in header]
        else:
            header_adjusted = header
        
        # è½¬æ¢ä¸º0åŸºå¹¶éªŒè¯è¾¹ç•Œ
        header_0_based = [h - 1 for h in header_adjusted if h > 0]
        header_0_based = [h for h in header_0_based if 0 <= h < df_length]
        
        if not header_0_based:
            logger.warning("    æœªæ‰¾åˆ°æœ‰æ•ˆçš„è¡¨å¤´è¡Œï¼Œä½¿ç”¨é»˜è®¤å€¼")
            return [0] if df_length > 0 else None
        
        return header_0_based
    
    def _extract_data_with_headers(self, df_raw: pd.DataFrame, header_0_based: List[int]) -> pd.DataFrame:
        """
        ä½¿ç”¨æŒ‡å®šçš„è¡¨å¤´è¡Œä»æ•°æ®æ¡†æå–æ•°æ®
        
        Args:
            df_raw: åŸå§‹æ•°æ®æ¡†ï¼ˆæ— è¡¨å¤´ï¼‰
            header_0_based: è¡¨å¤´è¡Œç´¢å¼•ï¼ˆ0åŸºï¼‰
            
        Returns:
            è®¾ç½®äº†è¡¨å¤´å¹¶æå–äº†æ•°æ®çš„DataFrame
        """
        if len(header_0_based) == 1:
            # å•çº§è¡¨å¤´
            header_idx = header_0_based[0]
            if header_idx >= len(df_raw):
                # è¡¨å¤´ç´¢å¼•è¶Šç•Œï¼Œä½¿ç”¨é»˜è®¤åˆ—å
                df = df_raw.copy()
                df.columns = [f'Column_{i}' for i in range(len(df.columns))]
                return df
            
            df_raw.columns = df_raw.iloc[header_idx]
            data_start = header_idx + 1
            
            if data_start < len(df_raw):
                return df_raw.iloc[data_start:].reset_index(drop=True)
            else:
                # è¡¨å¤´åæ²¡æœ‰æ•°æ®è¡Œ
                return pd.DataFrame(columns=df_raw.iloc[header_idx])
        else:
            # å¤šçº§è¡¨å¤´ï¼šåˆå¹¶ä¸ºå•è¡Œ
            header_rows_data = df_raw.iloc[header_0_based]
            new_columns = []
            
            for col_idx in range(len(df_raw.columns)):
                col_values = []
                for row_idx in range(len(header_0_based)):
                    val = header_rows_data.iloc[row_idx, col_idx]
                    if pd.notna(val) and str(val).strip() and 'Unnamed' not in str(val):
                        val_str = str(val).strip()
                        # åªåŒ…å«çŸ­å€¼ï¼ˆå¯èƒ½æ˜¯è¡¨å¤´ï¼Œä¸æ˜¯æ•°æ®ï¼‰
                        if len(val_str) <= 50:  # è¡¨å¤´é€šå¸¸å¾ˆçŸ­
                            col_values.append(val_str)
                
                # å»é‡ä½†ä¿æŒé¡ºåº
                col_values_dedup = list(OrderedDict.fromkeys(col_values))
                # åªæœ‰åœ¨æœ‰åˆç†çš„è¡¨å¤´å€¼æ—¶æ‰è¿æ¥
                if col_values_dedup and all(len(v) <= 30 for v in col_values_dedup):
                    col_name = '-'.join(col_values_dedup)
                elif col_values_dedup:
                    # ä½¿ç”¨æœ€çŸ­/æœ€å¯èƒ½çš„è¡¨å¤´å€¼
                    col_name = min(col_values_dedup, key=len)
                else:
                    col_name = f'Column_{col_idx}'
                new_columns.append(col_name)
            
            df_raw.columns = new_columns
            data_start = max(header_0_based) + 1
            return df_raw.iloc[data_start:].reset_index(drop=True)
    
    def _clean_column_names(self, columns: pd.Index) -> List[str]:
        """
        æ¸…ç†å’Œæ ‡å‡†åŒ–åˆ—å
        
        Args:
            columns: åŸå§‹åˆ—ç´¢å¼•
            
        Returns:
            æ¸…ç†åçš„åˆ—ååˆ—è¡¨
        """
        return [
            str(col).strip() if col and str(col).strip() else f'Column_{i}'
            for i, col in enumerate(columns)
        ]
    
    def _reshape_to_2d(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        å°†DataFrameé‡å¡‘ä¸ºæ ‡å‡†çš„äºŒç»´è¡¨
        
        Args:
            df: åŸå§‹DataFrame
            
        Returns:
            é‡å¡‘åçš„DataFrame
        """
        # åˆ é™¤å®Œå…¨ä¸ºç©ºçš„è¡Œå’Œåˆ—
        df = df.dropna(how='all').dropna(axis=1, how='all')
        
        # é‡ç½®ç´¢å¼•
        df = df.reset_index(drop=True)
        
        # æ¸…ç†åˆ—åï¼šå»é™¤ç©ºæ ¼ï¼Œå¤„ç†ç‰¹æ®Šå­—ç¬¦ï¼Œå¤„ç†NaNå€¼
        cleaned_columns = []
        for col in df.columns:
            if pd.isna(col):
                cleaned_columns.append(f"Unnamed_{len(cleaned_columns)}")
            else:
                cleaned_columns.append(str(col).strip())
        df.columns = cleaned_columns
        
        # ç¡®ä¿åˆ—åå”¯ä¸€
        df.columns = self._make_unique_columns(df.columns)
        
        return df
    
    def _make_unique_columns(self, columns: List) -> List:
        """ç¡®ä¿åˆ—åå”¯ä¸€"""
        seen = {}
        result = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                result.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                result.append(col)
        return result
    
    def load_all_files(self) -> Dict[str, pd.DataFrame]:
        """
        åŠ è½½çŸ¥è¯†åº“ä¸­çš„æ‰€æœ‰Excelæ–‡ä»¶
        
        Returns:
            æ–‡ä»¶ååˆ°DataFrameçš„æ˜ å°„
        """
        if not os.path.exists(self.knowledge_base_path):
            return {}
        
        excel_files = [f for f in os.listdir(self.knowledge_base_path) 
                      if f.endswith(('.xlsx', '.xls'))]
        
        for file in excel_files:
            file_path = os.path.join(self.knowledge_base_path, file)
            try:
                self.load_excel_file(file_path)
            except Exception as e:
                logger.warning(f"æ— æ³•åŠ è½½æ–‡ä»¶ {file}: {str(e)}")
        
        return self.processed_files
    
    def get_file_info(self, file_name: str) -> Optional[Dict]:
        """è·å–æ–‡ä»¶ä¿¡æ¯"""
        return self.file_metadata.get(file_name)
    
    def get_all_files_info(self) -> Dict[str, Dict]:
        """è·å–æ‰€æœ‰æ–‡ä»¶çš„ä¿¡æ¯"""
        return self.file_metadata
    
    def search_files_by_keywords(self, keywords: List[str]) -> List[str]:
        """
        æ ¹æ®å…³é”®è¯æœç´¢ç›¸å…³æ–‡ä»¶
        
        Args:
            keywords: å…³é”®è¯åˆ—è¡¨
            
        Returns:
            åŒ¹é…çš„æ–‡ä»¶ååˆ—è¡¨
        """
        matching_files = []
        keywords_lower = [k.lower() for k in keywords]
        
        for file_name, metadata in self.file_metadata.items():
            # æ£€æŸ¥æ–‡ä»¶å
            file_lower = file_name.lower()
            if any(kw in file_lower for kw in keywords_lower):
                matching_files.append(file_name)
                continue
            
            # æ£€æŸ¥åˆ—å
            columns = [str(col).lower() for col in metadata['columns']]
            if any(kw in ' '.join(columns) for kw in keywords_lower):
                matching_files.append(file_name)
        
        return matching_files
    
    def _print_preprocessed_headers(self, file_name: str, df: pd.DataFrame) -> None:
        """
        æ‰“å°é¢„å¤„ç†åçš„è¡¨å¤´ä¿¡æ¯
        
        Args:
            file_name: æ–‡ä»¶å
            df: é¢„å¤„ç†åçš„DataFrame
        """
        print("\n" + "=" * 70)
        print(f"ğŸ“Š é¢„å¤„ç†åçš„è¡¨å¤´ä¿¡æ¯: {file_name}")
        print("=" * 70)
        print(f"æ•°æ®å½¢çŠ¶: {df.shape[0]} è¡Œ Ã— {df.shape[1]} åˆ—")
        print(f"\nåˆ—ååˆ—è¡¨ ({len(df.columns)} åˆ—):")
        print("-" * 70)
        for i, col in enumerate(df.columns, 1):
            dtype = df[col].dtype
            non_null_count = df[col].notna().sum()
            print(f"  {i:2d}. {col:30s} | ç±»å‹: {str(dtype):10s} | éç©ºå€¼: {non_null_count}/{len(df)}")
        
        print("\nå‰5è¡Œæ•°æ®é¢„è§ˆ:")
        print("-" * 70)
        if len(df) > 0:
            # æ˜¾ç¤ºå‰5è¡Œï¼Œé™åˆ¶åˆ—å®½ä»¥ä¾¿æŸ¥çœ‹
            preview_df = df.head(5)
            # å¦‚æœåˆ—å¤ªå¤šï¼Œåªæ˜¾ç¤ºå‰10åˆ—
            if len(df.columns) > 10:
                preview_df = preview_df.iloc[:, :10]
                print(f"(ä»…æ˜¾ç¤ºå‰10åˆ—ï¼Œå…±{len(df.columns)}åˆ—)")
            print(preview_df.to_string())
        else:
            print("  (æ•°æ®ä¸ºç©º)")
        
        print("=" * 70 + "\n")
